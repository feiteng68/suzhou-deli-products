import openpyxl
import json

wb = openpyxl.load_workbook(r'image\产品图片规格.xlsx')
ws = wb.active

print(f"Sheet name: {ws.title}")
print(f"Dimensions: {ws.dimensions}")
print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
print("\nHeaders:")
for col in range(1, ws.max_column + 1):
    print(f"  Col {col}: {ws.cell(row=1, column=col).value}")

print("\nData rows:")
for row in range(2, min(ws.max_row + 1, 10)):
    row_data = []
    for col in range(1, ws.max_column + 1):
        row_data.append(ws.cell(row=row, column=col).value)
    print(f"  Row {row}: {row_data}")

products = []
for row in range(2, ws.max_row + 1):
    product = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        value = ws.cell(row=row, column=col).value
        if header:
            product[header] = value
    if product:
        products.append(product)

print(f"\nTotal products: {len(products)}")
print("\nJSON output:")
print(json.dumps(products, ensure_ascii=False, indent=2))

with open('products_data.json', 'w', encoding='utf-8') as f:
    json.dump(products, f, ensure_ascii=False, indent=2)

print("\nSaved to products_data.json")

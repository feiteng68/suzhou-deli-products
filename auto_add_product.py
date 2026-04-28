import os
import json
import shutil
from datetime import datetime

class ProductManager:
    def __init__(self, project_root):
        self.project_root = project_root
        self.image_dir = os.path.join(project_root, 'image')
        self.products_json = os.path.join(project_root, 'products.json')
        self.main_js = os.path.join(project_root, 'main.js')
        self.new_products_dir = os.path.join(project_root, 'new_products')
        self.backup_dir = os.path.join(project_root, 'backups')
        
        # 确保目录存在
        os.makedirs(self.new_products_dir, exist_ok=True)
        os.makedirs(self.backup_dir, exist_ok=True)
        
    def load_products(self):
        """加载现有产品数据"""
        if os.path.exists(self.products_json):
            with open(self.products_json, 'r', encoding='utf-8') as f:
                return json.load(f)
        return []
    
    def save_products(self, products):
        """保存产品数据（先备份）"""
        # 备份原文件
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = os.path.join(self.backup_dir, f'products_{timestamp}.json')
        if os.path.exists(self.products_json):
            shutil.copy2(self.products_json, backup_file)
            print(f"已备份原文件到: {backup_file}")
        
        # 保存新文件
        with open(self.products_json, 'w', encoding='utf-8') as f:
            json.dump(products, f, ensure_ascii=False, indent=2)
        print(f"产品数据已更新，共 {len(products)} 个产品")
    
    def scan_new_images(self):
        """扫描新图片"""
        new_images = []
        existing_products = self.load_products()
        existing_names = [p['产品名称'].replace(' ', '') for p in existing_products]
        
        # 扫描 image 目录
        for filename in os.listdir(self.image_dir):
            if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                product_name = os.path.splitext(filename)[0]
                if product_name not in existing_names:
                    new_images.append({
                        'filename': filename,
                        'product_name': product_name,
                        'path': os.path.join(self.image_dir, filename)
                    })
        
        # 扫描 new_products 目录
        if os.path.exists(self.new_products_dir):
            for filename in os.listdir(self.new_products_dir):
                if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
                    product_name = os.path.splitext(filename)[0]
                    if product_name not in existing_names:
                        new_images.append({
                            'filename': filename,
                            'product_name': product_name,
                            'path': os.path.join(self.new_products_dir, filename),
                            'is_new': True
                        })
        
        return new_images
    
    def read_product_info(self, product_name):
        """从 Excel 或文本文件读取产品信息"""
        info_file = os.path.join(self.new_products_dir, f'{product_name}.txt')
        if os.path.exists(info_file):
            with open(info_file, 'r', encoding='utf-8') as f:
                content = f.read()
                lines = content.strip().split('\n')
                info = {'specs': None, 'description': None}
                
                for line in lines:
                    if line.startswith('规格:'):
                        info['specs'] = line.replace('规格:', '').strip()
                    elif line.startswith('介绍:') or line.startswith('产品介绍:'):
                        info['description'] = line.split(':', 1)[1].strip()
                
                return info
        return {'specs': None, 'description': None}
    
    def get_pinyin_suggestion(self, product_name):
        """生成拼音建议"""
        pinyin_map = {
            'A': 'A', 'B': 'B', 'C': 'C', 'D': 'D', 'E': 'E', 'F': 'F',
            'G': 'G', 'H': 'H', 'I': 'I', 'J': 'J', 'K': 'K', 'L': 'L',
            'M': 'M', 'N': 'N', 'O': 'O', 'P': 'P', 'Q': 'Q', 'R': 'R',
            'S': 'S', 'T': 'T', 'U': 'U', 'V': 'V', 'W': 'W', 'X': 'X',
            'Y': 'Y', 'Z': 'Z',
            '0': '0', '1': '1', '2': '2', '3': '3', '4': '4',
            '5': '5', '6': '6', '7': '7', '8': '8', '9': '9'
        }
        
        # 如果是英文或数字开头，直接返回
        if product_name[0] in pinyin_map:
            return pinyin_map[product_name[0]]
        
        # 中文需要用户输入拼音
        return None
    
    def add_product(self, product_name, specs=None, description=None):
        """添加新产品"""
        products = self.load_products()
        
        # 检查是否已存在
        for p in products:
            if p['产品名称'].replace(' ', '') == product_name.replace(' ', ''):
                print(f"产品 '{product_name}' 已存在，跳过")
                return False
        
        new_product = {
            '产品名称': product_name,
            '规格': specs,
            '产品介绍': description
        }
        
        products.append(new_product)
        self.save_products(products)
        print(f"已添加产品: {product_name}")
        return True
    
    def update_pinyin_map(self, product_name, pinyin):
        """更新 main.js 中的拼音映射"""
        if not os.path.exists(self.main_js):
            print("main.js 不存在，跳过拼音更新")
            return
        
        with open(self.main_js, 'r', encoding='utf-8') as f:
            content = f.read()
        
        # 检查是否已存在
        if f"'{product_name}': '{pinyin}'" in content:
            print(f"拼音映射已存在: {product_name} -> {pinyin}")
            return
        
        # 找到 pinyinMap 对象并添加新映射
        import re
        pattern = r"(const pinyinMap = \{[^}]*?)(\};)"
        match = re.search(pattern, content, re.DOTALL)
        
        if match:
            pinyin_map_content = match.group(1)
            # 添加新的拼音映射（按字母顺序）
            new_entry = f"    '{product_name}': '{pinyin}',\n"
            updated_content = pinyin_map_content + new_entry + "  };"
            content = content.replace(match.group(0), updated_content)
            
            with open(self.main_js, 'w', encoding='utf-8') as f:
                f.write(content)
            print(f"已更新拼音映射: {product_name} -> {pinyin}")
        else:
            print("未找到 pinyinMap，请手动添加")
    
    def copy_new_images(self, new_images):
        """将新图片复制到 image 目录"""
        for img in new_images:
            if img.get('is_new'):
                dest_path = os.path.join(self.image_dir, img['filename'])
                if not os.path.exists(dest_path):
                    shutil.copy2(img['path'], dest_path)
                    print(f"已复制图片: {img['filename']}")
    
    def run_auto_add(self):
        """自动添加新产品"""
        print("=" * 50)
        print("苏州德力产品管理工具 - 自动添加")
        print("=" * 50)
        
        # 扫描新图片
        new_images = self.scan_new_images()
        
        if not new_images:
            print("未发现新图片")
            return
        
        print(f"\n发现 {len(new_images)} 个新产品图片:")
        for i, img in enumerate(new_images, 1):
            print(f"  {i}. {img['product_name']} ({img['filename']})")
        
        # 复制新图片
        self.copy_new_images(new_images)
        
        # 为每个新产品添加数据
        for img in new_images:
            product_name = img['product_name']
            print(f"\n处理产品: {product_name}")
            
            # 读取产品信息
            info = self.read_product_info(product_name)
            
            # 如果文本文件中有信息，直接使用
            if info['specs'] or info['description']:
                print(f"  从文本文件读取到信息")
                if info['specs']:
                    print(f"  规格: {info['specs'][:50]}...")
                if info['description']:
                    print(f"  介绍: {info['description'][:30]}...")
                
                self.add_product(product_name, info['specs'], info['description'])
                
                # 提示用户输入拼音
                pinyin = input(f"  请输入 '{product_name}' 的拼音首字母 (如 qingchuan): ")
                if pinyin:
                    self.update_pinyin_map(product_name, pinyin)
            else:
                # 交互式添加
                print(f"  未找到产品信息文件，请手动输入:")
                specs = input(f"  规格 (直接回车跳过): ").strip() or None
                description = input(f"  产品介绍 (直接回车跳过): ").strip() or None
                
                self.add_product(product_name, specs, description)
                
                # 提示输入拼音
                pinyin = input(f"  请输入 '{product_name}' 的拼音首字母 (如 qingchuan): ")
                if pinyin:
                    self.update_pinyin_map(product_name, pinyin)
        
        print("\n" + "=" * 50)
        print("处理完成！")
        print("请刷新页面查看新产品: http://localhost:8080/index.html")
        print("=" * 50)
    
    def run_batch_import(self, excel_file):
        """从 Excel 批量导入"""
        try:
            import openpyxl
        except ImportError:
            print("请先安装 openpyxl: pip install openpyxl")
            return
        
        print("=" * 50)
        print("苏州德力产品管理工具 - 批量导入")
        print("=" * 50)
        
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        products = []
        for row in range(2, ws.max_row + 1):
            name = ws.cell(row=row, column=1).value
            specs = ws.cell(row=row, column=2).value
            description = ws.cell(row=row, column=3).value
            
            if name:
                products.append({
                    '产品名称': str(name).strip(),
                    '规格': str(specs).strip() if specs else None,
                    '产品介绍': str(description).strip() if description else None
                })
        
        # 合并到现有数据
        existing = self.load_products()
        existing_names = [p['产品名称'] for p in existing]
        
        added_count = 0
        for product in products:
            if product['产品名称'] not in existing_names:
                existing.append(product)
                added_count += 1
        
        self.save_products(existing)
        print(f"成功导入 {added_count} 个新产品")
        
        # 提示更新拼音
        print("\n请记得在 main.js 中添加新产品的拼音映射！")

if __name__ == '__main__':
    import sys
    
    project_root = os.path.dirname(os.path.abspath(__file__))
    manager = ProductManager(project_root)
    
    if len(sys.argv) > 1:
        command = sys.argv[1]
        
        if command == 'auto':
            manager.run_auto_add()
        elif command == 'import' and len(sys.argv) > 2:
            manager.run_batch_import(sys.argv[2])
        elif command == 'help':
            print("使用方法:")
            print("  python auto_add_product.py auto          # 自动检测并添加新产品")
            print("  python auto_add_product.py import <文件>  # 从 Excel 批量导入")
            print("  python auto_add_product.py help           # 显示帮助")
        else:
            print("未知命令，使用 'help' 查看帮助")
    else:
        manager.run_auto_add()
